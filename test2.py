import openpyxl
import os
from pathlib import Path


parent_dir = os.path.join(os.getcwd(), "1", "Sample-Spreadsheet-50000-rows.xlsx")
# Define variable to load the dataframe

print(parent_dir)


dataframe = openpyxl.load_workbook(parent_dir)

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)
